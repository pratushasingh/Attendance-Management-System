# -*- coding: utf-8 -*-
"""
Created on Tue Apr 30 12:14:43 2019

@author: PRATUSHA
"""

# -*- coding: utf-8 -*-
"""

"""

from tkinter import *
import urllib
import cv2
import numpy as np
from imutils import paths
import face_recognition
import argparse
import pickle
import os
from openpyxl import Workbook
import datetime
from openpyxl import load_workbook
import os.path
 
window = Tk()
 
window.title("Face Recognition")
window.geometry('350x200')
 
lbl = Label(window, text="Enter your Roll No.")
 
lbl.grid(column=0, row=0)
txt = Entry(window,width=10)
 
txt.grid(column=1, row=0)
'''lbl = Label(window, text="Enter your Name")
 
lbl.grid(column=0, row=3)
txt1 = Entry(window,width=10)
 
txt1.grid(column=1, row=3)'''
#global sampleNum1
#sampleNum1=0
def clicked3():
    
    url="http://192.168.43.5:8080/shot.jpg"
    sampleNum1=0
    while(True):
        imgResp=urllib.request.urlopen(url)
        imgNp=np.array(bytearray(imgResp.read()),dtype=np.uint8)
        img=cv2.imdecode(imgNp,-1)
        cv2.imwrite("take_attendance/"+ str(sampleNum1) + ".jpg",img)
        sampleNum1+=1
        cv2.imshow('frame',img)
        if cv2.waitKey(10) & 0xFF == ord('q'):
            break
        elif sampleNum1>5:
            break
    cv2.destroyAllWindows()     
    
    
    
    
    
    
def clicked2():
    print("[INFO] loading encodings...")
    data = pickle.loads(open(("encodings4.pickle"), "rb").read())
    
    now= datetime.datetime.now()
    today=now.day
    month=now.month
    j=2
    title=['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30']
    #df=pd.DataFrame(columns=title)
    #writer=ExcelWriter(str(month)+'.xlsx')
    #df.to_excel()
    #wb = load_workbook('./'+str(month)+'.xlsx')
    #sheet=wb.active
    fname=(str(month)+'.xlsx')
    if os.path.isfile(fname):
     wb = load_workbook('./'+str(month)+'.xlsx')
    else:
     wb=Workbook()
    sheet=wb.active
     
    for i in title[0:30]:
     sheet.cell(row=1,column=j).value=i
     j+=1
     #sheet.cell(row=1,column=j).value=t
     #j+=1
    sheet.cell(row=3, column=1).value = "Arjun"
    sheet.cell(row=4, column=1).value = "Divyansh"
    sheet.cell(row=5, column=1).value = "Nazar" 
    sheet.cell(row=6, column=1).value = "Pratusha" 
    sheet.cell(row=7, column=1).value = "Priyanka" 
    sheet.cell(row=8, column=1).value = "Sagar" 
    sheet.cell(row=9, column=1).value = "Shakshi" 
    sheet.cell(row=10, column=1).value = "Shagun" 
    sheet.cell(row=11, column=1).value = "Shantanu" 
    sheet.cell(row=12, column=1).value = "Shivam" 
    sheet.cell(row=13, column=1).value = "Shreya" 
    sheet.cell(row=14, column=1).value = "Tehreem"
    sheet.cell(row=15, column=1).value = "Pratishtha"
    sheet.cell(row=16, column=1).value = "Aditi" 
    sheet.cell(row=17, column=1).value = "Anubha"
    sheet.cell(row=18, column=1).value = "Faran"
    sheet.cell(row=19, column=1).value = "Habeeba"
    
    
       

    image = cv2.imread("examplesdone1/example_60.jpg ")
    rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    
    print("[INFO] recognizing faces...")
    boxes = face_recognition.face_locations(rgb,
    	model="hog")
    encodings = face_recognition.face_encodings(rgb, boxes)
    
    
    names = []
    
    for encoding in encodings:
    	
     matches = face_recognition.compare_faces(data["encodings"],
    		encoding)
     name = "Unknown"
    
    	
     if True in matches:
    		
      matchedIdxs = [i for (i, b) in enumerate(matches) if b]
      counts = {}
    
    		
      for i in matchedIdxs:
       name = data["names"][i]
       counts[name] = counts.get(name, 0) + 1
    
    		
      name = max(counts, key=counts.get)
      sheet.cell(row=int(name)+2, column=int(today)+1).value = "Present"
     
    	
    	 
     names.append(name)
    
    
    for ((top, right, bottom, left), name) in zip(boxes, names):
    	
    	cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)
    	y = top - 15 if top - 15 > 15 else top + 15
    	cv2.putText(image, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX,
    		1.50, (0, 255, 0), 4)
    #path ='C:\Users\PRATUSHA\Downloads\Neq'
    #cv2.imwrite(os.path.join('C:\Users\PRATUSHA\Downloads\Neq','1.jpg'),image)
    image=cv2.resize(image,(1000,600))
    
    wb.save(str(month)+'.xlsx')
    cv2.imshow("Image", image)
    cv2.waitKey(0)
    			   
        
    
    
    

def clicked1():
    knownEncodings = []
    knownNames = []
    print("[INFO] quantifying faces...")
    imagePaths = list(paths.list_images("dataset_code"))


    for (i, imagePath) in enumerate(imagePaths):
    	
    	print("[INFO] processing image {}/{}".format(i + 1,
    		len(imagePaths)))
    	name = imagePath.split(os.path.sep)[-2]
    
    	
    	image = cv2.imread(imagePath)
    	rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    	
    	boxes = face_recognition.face_locations(rgb,
    		model="hog")
    
    	
    	encodings = face_recognition.face_encodings(rgb, boxes)
    
    	
    	for encoding in encodings:
    		
    		knownEncodings.append(encoding)
    		knownNames.append(name)
    

    print("[INFO] serializing encodings...")
    data = {"encodings": knownEncodings, "names": knownNames}
    f = open("encodings4.pickle", "wb")
    f.write(pickle.dumps(data))
    f.close()

    
    
    
  
def clicked():
 
    #lbl.configure(text="Button was clicked !!")
 url="http://192.168.43.5:8080/shot.jpg"
 detector=cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

# Id=input('enter your id')
 Id=txt.get()
 #res = "Welcome to " + txt.get()
 
 #lbl.configure(text= res)
 sampleNum=0
 while(True):
   # ret, img = cam.read()
  imgResp=urllib.request.urlopen(url)
 
  imgNp=np.array(bytearray(imgResp.read()),dtype=np.uint8)
  img=cv2.imdecode(imgNp,-1)
  #gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
  faces = detector.detectMultiScale(img, 1.3, 5)
  
  for (x,y,w,h) in faces:
   cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)
        
         
   sampleNum=sampleNum+1
        
   cv2.imwrite("dataset_code/" +str(Id)+'/'+ str(sampleNum) + ".jpg", img[y:y+h,x:x+w])

   cv2.imshow('frame',img)
     
  if cv2.waitKey(100) & 0xFF == ord('q'):
   break

  elif sampleNum>25:
   break
#cam.release()
 cv2.destroyAllWindows()
 #res=txt.get()
 #lbl.configure(text= res)

#cam = cv2.VideoCapture(1)
         
 
btn = Button(window, text="Take input", command=clicked)
 
btn.grid(column=0, row=5)
btn = Button(window, text="train images", command=clicked1)
 
btn.grid(column=1, row=5)
btn = Button(window, text="Quit", command=clicked2)

 
btn.grid(column=3, row=5)

btn = Button(window, text="Take attendance", command=clicked3)
btn.grid(column=2, row=5)
 
 
window.mainloop()